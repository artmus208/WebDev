from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Color

from app.report.reports_in_xl import autosize_columns
from app.report.reports_in_xl import autosize_columns_by_head
from app.reports_makers import project_report2


def create_xl_project(projects_data):

    wb = Workbook()
    ws = wb.active

    headers = [
        "Код проекта",
        "Название проекта",
        "Статьи расходов",
        "План, чел/день",
        "Факт, чел/день",
        "Отклонение, чел/день",
        "Отклонение, %"
    ]

    bold_font = Font(bold=True)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = bold_font
    row_number = 2
    for project_name, expenses in projects_data.items():

        for expense in expenses:
            col_num_ = 4
            ws.append([project_name] + expense)

            grey_bold_font = Font(bold=True, color="808080")
            red_bold_font = Font(bold=True, color="FF0000")
            green_bold_font = Font(bold=True, color="06a77d")
            black_bold_font = Font(bold=True, color="000000")

            for i in range(len(expense)):
                font_to_apply = bold_font

                if col_num_ == 4:
                    font_to_apply = grey_bold_font
                elif col_num_ == 5:
                    if expense[3] > expense[2]:
                        font_to_apply = red_bold_font
                    elif expense[3] < expense[2]:
                        font_to_apply = green_bold_font
                    else:
                        font_to_apply = black_bold_font
                elif col_num_ == 6:
                    if float(expense[4]) < 0:
                        font_to_apply = red_bold_font
                    else:
                        font_to_apply = black_bold_font

                elif col_num_ == 7:
                    if float(expense[5]) < 0:
                        font_to_apply = red_bold_font
                    else:
                        font_to_apply = black_bold_font

                ws.cell(row=row_number, column=col_num_).font = font_to_apply

                col_num_ += 1
            row_number += 1

    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 26 if col != 'B' else 45

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

        ws.column_dimensions[col].width = 30
        
    autosize_columns_by_head(ws)
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return file_stream


def get_info_project(id):
    detailed_p_report = project_report2(id)

    info_project = []

    c_c_list = detailed_p_report["cat_cost_list"]
    for i, c_c_name in enumerate(c_c_list):
        d_p_row = [
            c_c_name,
            round(c_c_list[c_c_name]["cat_cost_plan"] / 60 / 8, 1),
            round(c_c_list[c_c_name]["total_perf_time"] / 60 / 8, 1),
            round(c_c_list[c_c_name]["abs_diff"] / 60 / 8, 1),
            round(c_c_list[c_c_name]["rel_diff"], 1),
        ]
        info_project.append(d_p_row)
    return info_project

