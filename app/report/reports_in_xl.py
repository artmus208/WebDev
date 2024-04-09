# Здесь я выгружаю пока что только краткий отчет за всё время
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from app.reports_makers import project_report2

from app import app

def autosize_columns(worksheet):
    def value_of(value):
        return (str(value) if value is not None else "")

    for cells in worksheet.columns:
        length = max(len(value_of(cell.value)) for cell in cells)
        column_letter = get_column_letter(cells[0].column)
        worksheet.column_dimensions[column_letter].width = length

    return worksheet

# TIPS: Эту функцию надо рефакторить обязательно
def brief_p_report_xl(p_id=14):

    # TODO:
    # [ ]: Вынести Workbook в аргументы
    
    detailed_p_report = project_report2(p_id)

    plan_labor_costs = f"Планируемые трудозатраты: {detailed_p_report['plan_time'] // 60 // 8 } чел/д"
    actual_labor_costs = f"Фактические трудозатраты: {detailed_p_report['total_perf_time'] // 60 // 8 } чел/д"
    abs_diff = f"Разница: {detailed_p_report['abs_diff'] // 60 // 8 } чел/д"
    interest = f"Разница в %: {detailed_p_report['rel_diff']} %"

    values_for_column_b = [detailed_p_report['plan_time'] // 60 // 8, detailed_p_report['total_perf_time'] // 60 // 8, detailed_p_report['abs_diff'] // 60 // 8, detailed_p_report['rel_diff']]


    p_title = detailed_p_report["p_name"]
    p_code = p_title.split()[0]

    wb = Workbook()

    ws: Worksheet = wb.create_sheet(f"Сжатый отчет {p_code}", 0)

    count_cat_costs = len(detailed_p_report["cat_cost_list"])
    count_rows = count_cat_costs + 4 + 1 # 4 - потому что таблица не от верхней границы, а от 4 строки вниз, 1 потмоу что есть ещё заголовок

    ref = ["E5", f"I{count_rows}"]

    caption_range = ws["E5":"I5"][0]

    for cell in caption_range:
        cell.font = Font(bold=True)

    grey_font = Font(color="808080")
    red_font = Font(color="FF0000")
    green_font = Font(color="00B800")

    caption_list = [
        "Статьи расходов", 
        "План, чел/день", 
        "Факт, чел/день",
        "Отклонение, чел/д", 
        "Отклонение, %", 
    ]

    for cell, caption in zip(caption_range, caption_list):
        cell.value = caption
    
    index_c_col = column_index_from_string("E")
    index_g_col = column_index_from_string("I")
    c_c_list = detailed_p_report["cat_cost_list"]
    for i, c_c_name in enumerate(c_c_list):
        d_p_row = [
        c_c_name,
        round(c_c_list[c_c_name]["cat_cost_plan"] / 60 / 8, 1),
        round(c_c_list[c_c_name]["total_perf_time"] / 60 / 8, 1),
        round(c_c_list[c_c_name]["abs_diff"] / 60 / 8, 1),
        round(c_c_list[c_c_name]["rel_diff"], 1),
    ]
        for row in ws.iter_rows(min_row=6+i, max_row=6+i, min_col=index_c_col, max_col=index_g_col):
            print(d_p_row)
            index = 0
            for cell, value in zip(row, d_p_row):
                cell.value = value

                if index == 1:
                    cell.font = Font(color="808080", bold=True)
                elif index == 2:
                    if d_p_row[2] > d_p_row[1]:
                        cell.font = Font(color="FF0000", bold=True)
                    else:
                        cell.font = Font(color="06a77d", bold=True)
                elif index == 3:
                    if d_p_row[3] < 0:
                        cell.font = Font(color="FF0000", bold=True)
                elif index == 4:
                    if d_p_row[4] < 0:
                        cell.font = Font(color="FF0000", bold=True)

                index += 1

    tab = Table(displayName="brief_project_report", ref=":".join(ref))
    style = TableStyleInfo(name="BriefStyle", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    autosize_columns(ws)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 20

    info = [f"Сжатый отчет по проекту", f"{p_title}", f"до {datetime.now().strftime('%d.%m.%Y')} включительно"]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=1)):
        for cell in row:
            cell.value = info[i]
            if i == 1:
                cell.font = Font(bold=True)

    new_rows_values = ["Плаинируемые трудозатраты:", "Фактические трудозатраты:", "Разница:", "Разница в %:"]
    for i, value in enumerate(new_rows_values, start=5):  # Начинаем с 5-й строки
        ws[f"A{i}"] = value

    for i, value in enumerate(values_for_column_b, start=5):  # Начинаем с 5-й строки
        ws[f"B{i}"].value = value
        ws[f"C{i}"].value = "чел/д" if i != 8 else "%"

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    wb.close()
    return file_stream, p_code

def write_emps_dict(ws: Worksheet, emps_dict: dict):
    """Функционал записи общего списка сотрудников по трудозатратам в этом проекте"""
    pass

def write_p_common_info(ws: Worksheet, p_common_info):
    """Функционал записи названия, общих показателей"""
    pass


def in_emps_dict(emp_data: dict):
    """Функционал составления общего списка сотрудников и их трудозатрат"""
    pass

def write_labor_table(ws: Worksheet, start_coord="E5", cc_name:str="Проектирование", cc_data:dict={}):
    """Функция записи таблицы с трудозатратами сотрудников по статьям расходов"""
    t_captions = ["Сотрудник", "Факт, чел/д"]
    
    emp_list = cc_data["emp_list"]
    
    min_row = int(start_coord[1])                      # Номер строки с названием статьи расходов
    min_col = column_index_from_string(start_coord[0]) # Номер колонки с названием статьи расходов
    max_row = 3 + len(emp_list) + min_row-1            # Номер строки для итого
    max_col = min_col + 1                              # Номер колонки последней ячейки в итого
    
    iter_rows = ws.iter_rows(min_row=min_row, max_row=max_row-1, min_col=min_col, max_col=max_col)
    cc_row = next(iter_rows)
    cc_row[0].value = cc_name
    
    caption_row = next(iter_rows)
    caption_row[0].value, caption_row[1].value = t_captions
    
    for row, emp in zip(iter_rows, emp_list):
        row[0].value = emp
        row[1].value = round(emp_list[emp]["total_perf_time"][0] / 60 / 8, 1)
        
                
    # TODO
    # Оформить её как table
    return max_row


def week_create_xl(projects_data, caption,):
    dict_pd = projects_data

    name_report = caption.split(",")

    p_title = name_report[0]
    p_code = p_title.split()[0]

    wb = Workbook()

    ws: Worksheet = wb.create_sheet(f"Сжатый отчет {p_code}", 0)

    count_cat_costs = len(projects_data)
    count_rows = count_cat_costs + 4 + 1  # 4 - потому что таблица не от верхней границы, а от 4 строки вниз, 1 потмоу что есть ещё заголовок

    ref = ["D5", f"I{count_rows}"]

    caption_range = ws["D5":"I5"][0]

    for cell in caption_range:
        cell.font = Font(bold=True)

    grey_font = Font(color="808080")
    red_font = Font(color="FF0000")
    green_font = Font(color="00B800")

    caption_list = [
        "Статьи расходов",
        "План, чел/д",
        "Факт, чел/д",
        "Отклонение, чел/д",
        "Отклонение(отн)",
        "За неделю, чел/д"
    ]

    for cell, caption in zip(caption_range, caption_list):
        cell.value = caption

    index_c_col = column_index_from_string("D")
    index_g_col = column_index_from_string("I")

    for i, c_c_name in enumerate(dict_pd):

        val_list = [c_c_name, dict_pd[c_c_name]['plan_labor'], dict_pd[c_c_name]['fact_labor'],
                    dict_pd[c_c_name]['delta'], str(dict_pd[c_c_name]['progress']) + "%",
                    dict_pd[c_c_name]['week_labor']]

        for row in ws.iter_rows(min_row=6 + i, max_row=6 + i, min_col=index_c_col, max_col=index_g_col):
            index = 0
            for cell, value in zip(row, val_list):
                cell.value = value
                if index == 1:
                    cell.font = Font(bold=True, color="808080")
                elif index == 2:
                    if dict_pd[c_c_name]['fact_labor'] > dict_pd[c_c_name]['plan_labor']:
                        cell.font = Font(bold=True, color="FF0000")
                    else:
                        cell.font = Font(bold=True, color="06a77d")
                elif index == 3:
                    if dict_pd[c_c_name]['delta'] < 0:
                        cell.font = Font(bold=True, color="FF0000")
                elif index == 4:
                    if dict_pd[c_c_name]['progress'] < 0:
                        cell.font = Font(bold=True, color="FF0000")
                cell.alignment = Alignment(horizontal="right")

                index += 1


    tab = Table(displayName="brief_project_report", ref=":".join(ref))
    style = TableStyleInfo(name="BriefStyle", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    autosize_columns(ws)
    for col in ['A', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 25 if col != 'A' else 45

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

    info = [f"{p_title}", f"Сжатый {name_report[1]}", f"{name_report[-1]}"]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=1)):
        for cell in row:
            cell.value = info[i]
            if i == 1:
                cell.font = Font(bold=True)

    # new_rows_values = ["Плаинируемые трудозатраты:", "Фактические трудозатраты:", "Разница:", "Разница в %:"]
    # for i, value in enumerate(new_rows_values, start=5):  # Начинаем с 5-й строки
    #     ws[f"A{i}"] = value
    #
    # for i, value in enumerate(values_for_column_b, start=5):  # Начинаем с 5-й строки
    #     ws[f"B{i}"].value = value
    #     ws[f"C{i}"].value = "чел/д" if i != 8 else "%"

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    wb.close()
    return file_stream, p_code


# with app.app_context():
#     p_detail_report = project_report2(p_id=14)

# p_title = p_detail_report["p_name"]
# p_code = p_title.split()[0]

# wb = Workbook()
# ws = wb.create_sheet(f"Подробный отчет {p_code}")

# cc_name = list(p_detail_report["cat_cost_list"].keys())[0]
# cc_data = p_detail_report["cat_cost_list"][cc_name]
# write_labor_table(ws=ws, cc_name=cc_name, cc_data=cc_data)

# wb.save("test_detail_report.xlsx")


    # thin_border = Border(left=Side(style='thin', color=Color('4786ff')),
    #                      right=Side(style='thin', color=Color('4786ff')),
    #                      top=Side(style='thin', color=Color('4786ff')),
    #                      bottom=Side(style='thin', color=Color('4786ff')))
    #
    # # Применение границ к каждой ячейке в диапазоне таблицы
    # for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
    #     for cell in row:
    #         cell.border = thin_border